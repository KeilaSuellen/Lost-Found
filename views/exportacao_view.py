import customtkinter as ctk
from tkinter import messagebox, filedialog
import os, sys, threading, datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from database.db import listar_objetos, listar_retiradas


class ExportacaoView(ctk.CTkFrame):
    """Tela para exportar todos os registros do sistema em Excel ou PDF."""

    def __init__(self, master, usuario: dict, on_concluido=None):
        super().__init__(master, fg_color="transparent")
        self.usuario      = usuario
        self.on_concluido = on_concluido
        self._build_ui()

    def _build_ui(self):
        self.pack(expand=True, fill="both")

        ctk.CTkLabel(self, text="Exportar Registros",
                     font=ctk.CTkFont(size=20, weight="bold"),
                     ).pack(pady=(24, 4), anchor="w", padx=32)
        ctk.CTkLabel(self, text="Exporta objetos e histórico de retiradas em um único arquivo.",
                     font=ctk.CTkFont(size=12), text_color="gray",
                     ).pack(anchor="w", padx=32, pady=(0, 24))

        # ── Opções ────────────────────────────────────────────────────────────
        card = ctk.CTkFrame(self, corner_radius=10)
        card.pack(fill="x", padx=32, pady=(0, 16))

        ctk.CTkLabel(card, text="Filtro de objetos:",
                     font=ctk.CTkFont(size=13, weight="bold"), anchor="w",
                     ).pack(anchor="w", padx=16, pady=(14, 6))

        self.var_filtro = ctk.StringVar(value="todos")
        opts = [("Todos os objetos", "todos"),
                ("Apenas disponíveis", "disponivel"),
                ("Apenas retirados",  "retirado"),
                ("Apenas arquivados", "arquivado")]
        for label, val in opts:
            ctk.CTkRadioButton(card, text=label, variable=self.var_filtro,
                               value=val, font=ctk.CTkFont(size=13),
                               ).pack(anchor="w", padx=32, pady=2)

        ctk.CTkLabel(card, text="Incluir:", font=ctk.CTkFont(size=13, weight="bold"),
                     anchor="w").pack(anchor="w", padx=16, pady=(12, 6))

        self.var_retiradas = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(card, text="Histórico de retiradas",
                        variable=self.var_retiradas,
                        font=ctk.CTkFont(size=13),
                        ).pack(anchor="w", padx=32, pady=(0, 14))

        # ── Formato ───────────────────────────────────────────────────────────
        fmt_frame = ctk.CTkFrame(self, fg_color="transparent")
        fmt_frame.pack(fill="x", padx=32, pady=(0, 20))

        ctk.CTkButton(fmt_frame, text="Exportar Excel (.xlsx)", width=220, height=44,
                      command=self._exportar_excel).pack(side="left", padx=(0, 12))
        ctk.CTkButton(fmt_frame, text="Exportar PDF", width=220, height=44,
                      fg_color="#2a5a3a", hover_color="#1e4a2e",
                      command=self._exportar_pdf).pack(side="left")

        # ── Log ───────────────────────────────────────────────────────────────
        self.log = ctk.CTkTextbox(self, height=120, state="disabled")
        self.log.pack(fill="x", padx=32, pady=(0, 8))

        self.progress = ctk.CTkProgressBar(self)
        self.progress.pack(fill="x", padx=32)
        self.progress.set(0)

        bar = ctk.CTkFrame(self, fg_color="transparent")
        bar.pack(fill="x", padx=32, pady=16)
        ctk.CTkButton(bar, text="Voltar", width=120, fg_color="transparent",
                      border_width=1, command=self._voltar).pack(side="right")

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _log(self, txt):
        self.log.configure(state="normal")
        self.log.insert("end", txt + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def _dados(self):
        filtro = self.var_filtro.get()
        objetos = listar_objetos()
        if filtro != "todos":
            objetos = [o for o in objetos if o["status"] == filtro]
        retiradas = listar_retiradas() if self.var_retiradas.get() else []
        return objetos, retiradas

    def _nome_arquivo(self, ext):
        hoje = datetime.date.today().strftime("%Y%m%d")
        return f"achados_perdidos_{hoje}.{ext}"

    # ── Excel ─────────────────────────────────────────────────────────────────

    def _exportar_excel(self):
        path = filedialog.asksaveasfilename(
            title="Salvar Excel",
            defaultextension=".xlsx",
            initialfile=self._nome_arquivo("xlsx"),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        self.progress.set(0.1)
        threading.Thread(target=self._gerar_excel, args=(path,), daemon=True).start()

    def _gerar_excel(self, path):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side, GradientFill)
            from openpyxl.utils import get_column_letter

            objetos, retiradas = self._dados()
            self.after(0, lambda: self._log(f"Carregados {len(objetos)} objetos e {len(retiradas)} retiradas…"))
            self.after(0, lambda: self.progress.set(0.3))

            wb = Workbook()

            # ── Paleta ────────────────────────────────────────────────────────
            COR_HEADER  = "1a7a6e"   # verde escuro
            COR_TITULO  = "f0f8f6"   # verde muito claro
            COR_IMPAR   = "FFFFFF"
            COR_PAR     = "f4faf8"
            COR_DISP    = "d4edda"
            COR_RETIRADO= "f8d7da"
            COR_ARQUIVO = "fff3cd"

            borda = Border(
                left=Side(style="thin", color="cccccc"),
                right=Side(style="thin", color="cccccc"),
                top=Side(style="thin", color="cccccc"),
                bottom=Side(style="thin", color="cccccc"),
            )

            def estilar_header(ws, colunas, row=1):
                for col_idx, titulo in enumerate(colunas, 1):
                    cel = ws.cell(row=row, column=col_idx, value=titulo)
                    cel.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                    cel.fill      = PatternFill("solid", fgColor=COR_HEADER)
                    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cel.border    = borda

            def estilar_linha(ws, row, num_cols, par=False, cor_bg=None):
                bg = cor_bg or (COR_PAR if par else COR_IMPAR)
                for col in range(1, num_cols + 1):
                    cel = ws.cell(row=row, column=col)
                    cel.fill      = PatternFill("solid", fgColor=bg)
                    cel.alignment = Alignment(vertical="center", wrap_text=True)
                    cel.border    = borda
                    cel.font      = Font(name="Arial", size=10)

            # ABA 1 — Objetos
            ws_obj = wb.active
            ws_obj.title = "Objetos"
            ws_obj.row_dimensions[1].height = 14
            ws_obj.row_dimensions[2].height = 30

            # Título da aba
            ws_obj.merge_cells("A1:I1")
            t = ws_obj["A1"]
            t.value     = "ACHADOS E PERDIDOS — Relatório de Objetos"
            t.font      = Font(bold=True, size=13, color=COR_HEADER, name="Arial")
            t.fill      = PatternFill("solid", fgColor=COR_TITULO)
            t.alignment = Alignment(horizontal="center", vertical="center")

            # Subtítulo com data
            ws_obj.merge_cells("A2:I2")  
            s = ws_obj["A2"]
            s.value     = f"Gerado em {datetime.datetime.now().strftime('%d/%m/%Y às %H:%M')}"
            s.font      = Font(italic=True, size=10, color="888888", name="Arial")
            s.fill      = PatternFill("solid", fgColor=COR_TITULO)
            s.alignment = Alignment(horizontal="center", vertical="center")

            cols_obj = ["ID", "Nome", "Descrição", "Data Encontrada",
                        "Local Encontrado", "Quem Encontrou", "Empresa", "Status"]
            estilar_header(ws_obj, cols_obj, row=3)

            status_cores = {"disponivel": COR_DISP, "retirado": COR_RETIRADO, "arquivado": COR_ARQUIVO}
            status_label = {"disponivel": "Disponível", "retirado": "Retirado", "arquivado": "Arquivado"}

            for i, obj in enumerate(objetos):
                row = i + 4
                st  = obj.get("status", "disponivel")
                cor = status_cores.get(st, COR_IMPAR)
                ws_obj.cell(row=row, column=1, value=obj.get("id", ""))
                ws_obj.cell(row=row, column=2, value=obj.get("nome", ""))
                ws_obj.cell(row=row, column=3, value=obj.get("descricao", ""))
                ws_obj.cell(row=row, column=4, value=obj.get("data_encontrada", ""))
                ws_obj.cell(row=row, column=5, value=obj.get("local_encontrado", ""))
                ws_obj.cell(row=row, column=6, value=obj.get("quem_encontrou", ""))
                ws_obj.cell(row=row, column=7, value=obj.get("empresa", ""))
                ws_obj.cell(row=row, column=8, value=status_label.get(st, st))
                estilar_linha(ws_obj, row, 8, par=(i % 2 == 1), cor_bg=cor)

            # Totais
            tot_row = len(objetos) + 4
            ws_obj.merge_cells(f"A{tot_row}:G{tot_row}")
            tc = ws_obj.cell(row=tot_row, column=1, value=f"Total: {len(objetos)} objetos")
            tc.font      = Font(bold=True, name="Arial", size=10)
            tc.fill      = PatternFill("solid", fgColor=COR_TITULO)
            tc.alignment = Alignment(horizontal="right")
            tc.border    = borda
            ws_obj.cell(row=tot_row, column=8).fill   = PatternFill("solid", fgColor=COR_TITULO)
            ws_obj.cell(row=tot_row, column=8).border = borda

            # Larguras
            for col, w in zip("ABCDEFGH", [12, 22, 32, 16, 18, 20, 20, 16]):
                ws_obj.column_dimensions[col].width = w
            ws_obj.freeze_panes = "A4"

            # ABA 2 — Retiradas
            if retiradas:
                ws_ret = wb.create_sheet("Retiradas")
                ws_ret.row_dimensions[1].height = 14
                ws_ret.row_dimensions[2].height = 30

                ws_ret.merge_cells("A1:G1")
                t2 = ws_ret["A1"]
                t2.value     = "ACHADOS E PERDIDOS — Histórico de Retiradas"
                t2.font      = Font(bold=True, size=13, color=COR_HEADER, name="Arial")
                t2.fill      = PatternFill("solid", fgColor=COR_TITULO)
                t2.alignment = Alignment(horizontal="center", vertical="center")

                ws_ret.merge_cells("A2:G2")
                s2 = ws_ret["A2"]
                s2.value     = f"Gerado em {datetime.datetime.now().strftime('%d/%m/%Y às %H:%M')}"
                s2.font      = Font(italic=True, size=10, color="888888", name="Arial")
                s2.fill      = PatternFill("solid", fgColor=COR_TITULO)
                s2.alignment = Alignment(horizontal="center", vertical="center")

                cols_ret = ["ID Objeto", "Objeto", "Retirante",
                            "Documento", "Empresa", "Data Retirada"]
                estilar_header(ws_ret, cols_ret, row=3)

                for i, r in enumerate(retiradas):
                    row = i + 4
                    ws_ret.cell(row=row, column=1, value=r.get("id_objeto", ""))
                    ws_ret.cell(row=row, column=2, value=r.get("nome_objeto", ""))
                    ws_ret.cell(row=row, column=3, value=r.get("nome_retirante", ""))
                    ws_ret.cell(row=row, column=4, value=r.get("documento", ""))
                    ws_ret.cell(row=row, column=5, value=r.get("empresa", ""))
                    ws_ret.cell(row=row, column=6, value=r.get("data_retirada", ""))
                    estilar_linha(ws_ret, row, 6, par=(i % 2 == 1))

                tot_r = len(retiradas) + 4
                ws_ret.merge_cells(f"A{tot_r}:E{tot_r}")
                tr = ws_ret.cell(row=tot_r, column=1, value=f"Total: {len(retiradas)} retiradas")
                tr.font      = Font(bold=True, name="Arial", size=10)
                tr.fill      = PatternFill("solid", fgColor=COR_TITULO)
                tr.alignment = Alignment(horizontal="right")
                tr.border    = borda
                ws_ret.cell(row=tot_r, column=6).fill   = PatternFill("solid", fgColor=COR_TITULO)
                ws_ret.cell(row=tot_r, column=6).border = borda

                for col, w in zip("ABCDEF", [12, 22, 22, 16, 20, 16]):
                    ws_ret.column_dimensions[col].width = w
                ws_ret.freeze_panes = "A4"

            self.after(0, lambda: self.progress.set(0.8))
            wb.save(path)
            self.after(0, lambda: self.progress.set(1.0))
            self.after(0, lambda: self._log(f"✅ Excel salvo em:\n   {path}"))

            import subprocess, platform
            if platform.system() == "Windows":
                os.startfile(path)
            elif platform.system() == "Darwin":
                subprocess.call(["open", path])
            else:
                subprocess.call(["xdg-open", path])

        except ImportError:
            self.after(0, lambda: self._log(
                "❌ openpyxl não instalado.\nExecute: pip install openpyxl"))
        except Exception as e:
            self.after(0, lambda: self._log(f"❌ Erro: {e}"))

    # PDF 

    def _exportar_pdf(self):
        path = filedialog.asksaveasfilename(
            title="Salvar PDF",
            defaultextension=".pdf",
            initialfile=self._nome_arquivo("pdf"),
            filetypes=[("PDF", "*.pdf")],
        )
        if not path:
            return
        self.progress.set(0.1)
        threading.Thread(target=self._gerar_pdf, args=(path,), daemon=True).start()

    def _gerar_pdf(self, path):
        try:
            objetos, retiradas = self._dados()
            self.after(0, lambda: self._log(f"Gerando PDF com {len(objetos)} objetos…"))
            self.after(0, lambda: self.progress.set(0.3))

            status_label = {"disponivel": "Disponível", "retirado": "Retirado", "arquivado": "Arquivado"}
            status_cor   = {"disponivel": "#d4edda", "retirado": "#f8d7da", "arquivado": "#fff3cd"}

            def linhas_obj():
                rows = ""
                for obj in objetos:
                    st  = obj.get("status", "disponivel")
                    cor = status_cor.get(st, "#fff")
                    lbl = status_label.get(st, st)
                    rows += f"""<tr style="background:{cor}">
                        <td>{obj.get('id','')}</td>
                        <td>{obj.get('nome','')}</td>
                        <td>{obj.get('descricao','') or '-'}</td>
                        <td>{obj.get('data_encontrada','') or '-'}</td>
                        <td>{obj.get('local_encontrado','') or '-'}</td>
                        <td>{obj.get('quem_encontrou','') or '-'}</td>
                        <td>{obj.get('empresa','') or '-'}</td>
                        <td><strong>{lbl}</strong></td>
                    </tr>"""
                return rows

            def linhas_ret():
                if not retiradas:
                    return ""
                rows = ""
                for i, r in enumerate(retiradas):
                    bg = "#f4faf8" if i % 2 else "#ffffff"
                    rows += f"""<tr style="background:{bg}">
                        <td>{r.get('id_objeto','')}</td>
                        <td>{r.get('nome_objeto','')}</td>
                        <td>{r.get('nome_retirante','')}</td>
                        <td>{r.get('documento','') or '-'}</td>
                        <td>{r.get('empresa','') or '-'}</td>
                        <td>{r.get('data_retirada','') or '-'}</td>
                    </tr>"""
                return rows

            secao_ret = ""
            if retiradas:
                secao_ret = f"""
                <div class="secao">
                  <h2>Histórico de Retiradas <span class="badge">{len(retiradas)}</span></h2>
                  <table>
                    <thead><tr>
                      <th>ID Obj.</th><th>Objeto</th><th>Retirante</th>
                      <th>Documento</th><th>Empresa</th><th>Data Retirada</th>
                    </tr></thead>
                    <tbody>{linhas_ret()}</tbody>
                  </table>
                </div>"""

            agora = datetime.datetime.now().strftime("%d/%m/%Y às %H:%M")
            html  = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8">
<title>Relatório — Achados e Perdidos</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family: Arial, sans-serif; color: #1a2a28; background: #f0f4f4; padding: 32px; }}
  .capa {{
    background: linear-gradient(135deg, #1a7a6e, #0f4f46);
    color: white; border-radius: 12px;
    padding: 32px 40px; margin-bottom: 28px;
    display: flex; justify-content: space-between; align-items: center;
  }}
  .capa h1 {{ font-size: 24px; letter-spacing: 1px; }}
  .capa p  {{ font-size: 12px; opacity: 0.8; margin-top: 4px; }}
  .resumo {{
    display: flex; gap: 16px; margin-bottom: 28px;
  }}
  .card-resumo {{
    flex: 1; background: white; border-radius: 10px;
    padding: 16px 20px; text-align: center;
    border-top: 4px solid #1a7a6e;
    box-shadow: 0 1px 6px rgba(0,0,0,0.06);
  }}
  .card-resumo .num {{ font-size: 28px; font-weight: bold; color: #1a7a6e; }}
  .card-resumo .lbl {{ font-size: 11px; color: #666; margin-top: 4px; }}
  .secao {{
    background: white; border-radius: 10px;
    padding: 20px 24px; margin-bottom: 20px;
    box-shadow: 0 1px 6px rgba(0,0,0,0.06);
  }}
  .secao h2 {{
    font-size: 15px; color: #1a7a6e; margin-bottom: 14px;
    padding-bottom: 8px; border-bottom: 2px solid #e0eae8;
    display: flex; align-items: center; gap: 8px;
  }}
  .badge {{
    background: #1a7a6e; color: white;
    font-size: 11px; padding: 2px 8px; border-radius: 10px;
  }}
  table {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
  thead tr {{ background: #1a7a6e; color: white; }}
  th {{ padding: 8px 10px; text-align: left; font-weight: bold; }}
  td {{ padding: 7px 10px; border-bottom: 1px solid #e8f0ee; }}
  .rodape {{ text-align: center; font-size: 10px; color: #aaa; margin-top: 20px; }}
  @media print {{
    body {{ padding: 16px; background: white; }}
    .capa {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    thead tr {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    .no-print {{ display: none !important; }}
  }}
</style></head><body>

<div class="capa">
  <div>
    <h1>Achados e Perdidos</h1>
    <p>Relatório completo — gerado em {agora}</p>
  </div>
  <div style="text-align:right;font-size:13px;opacity:0.85">
    Verzani &amp; Sandrini / CBRE / EZ Towers
  </div>
</div>

<div class="resumo">
  <div class="card-resumo">
    <div class="num">{len(objetos)}</div>
    <div class="lbl">Objetos no relatório</div>
  </div>
  <div class="card-resumo">
    <div class="num">{sum(1 for o in objetos if o.get('status')=='disponivel')}</div>
    <div class="lbl">Disponíveis</div>
  </div>
  <div class="card-resumo">
    <div class="num">{sum(1 for o in objetos if o.get('status')=='retirado')}</div>
    <div class="lbl">Retirados</div>
  </div>
  <div class="card-resumo">
    <div class="num">{sum(1 for o in objetos if o.get('status')=='arquivado')}</div>
    <div class="lbl">Arquivados</div>
  </div>
  <div class="card-resumo">
    <div class="num">{len(retiradas)}</div>
    <div class="lbl">Retiradas registradas</div>
  </div>
</div>

<div class="secao">
  <h2>Objetos <span class="badge">{len(objetos)}</span></h2>
  <table>
    <thead><tr>
      <th>ID</th><th>Nome</th><th>Descrição</th><th>Data</th>
      <th>Local</th><th>Encontrado por</th><th>Empresa</th><th>Status</th>
    </tr></thead>
    <tbody>{linhas_obj()}</tbody>
  </table>
</div>

{secao_ret}

<div class="rodape">Achados e Perdidos — {agora}</div>

<div class="no-print" style="text-align:center;margin-top:28px">
  <button onclick="window.print()"
    style="padding:10px 36px;font-size:14px;cursor:pointer;
           background:#1a7a6e;color:white;border:none;border-radius:8px;font-weight:600">
        Imprimir / Salvar PDF
  </button>
</div>
</body></html>"""

            # Salva HTML temporário e abre no navegador para imprimir como PDF
            import tempfile, webbrowser
            tmp = tempfile.NamedTemporaryFile(
                delete=False, suffix=".html", mode="w", encoding="utf-8",
                prefix="relatorio_achados_"
            )
            tmp.write(html)
            tmp.close()
            webbrowser.open(f"file:///{tmp.name}")

            self.after(0, lambda: self.progress.set(1.0))
            self.after(0, lambda: self._log(
                "   Relatório aberto no navegador.\n"
                "   Use  Ctrl+P → Salvar como PDF  para salvar o arquivo."
            ))

        except Exception as e:
            self.after(0, lambda: self._log(f"❌ Erro: {e}"))

    def _voltar(self):
        if self.on_concluido:
            self.on_concluido()
